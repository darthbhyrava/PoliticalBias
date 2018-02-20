# -*- coding: utf-8 -*-
import nltk
import re
import openpyxl
from xml.dom import minidom
import xml.etree.ElementTree as ET

def dict_to_xml(tag, d):
    elem = ET.Element(tag)
    for key, val in d.items():
        child = ET.Element(key)
        child.text = str(val)
        elem.append(child)
    return elem

wb = openpyxl.load_workbook('./Corpus/37.xlsx')
ws = wb.active

lang_column = ws['F']
text_column = ws['J']
user_name_column = ws['T']
rows = len(text_column)

index = {}
users = {}
i = 0
for row in range(1,rows):
	text = text_column[row].value
	user = user_name_column[row].value
	lang = lang_column[row].value
	if lang == 'en':
		if user in users:
			users[user].append(text)
		else:
			users[user] = [text]
			index[user] = i
			i+=1

for key, value in users.items():
	num = index[key]
	value = '\n'.join(value)
	sentences = nltk.sent_tokenize(value)
	with open('./Results/{0}.xml'.format(num),'a+') as f:
		for ind, sentence in enumerate(sentences):
			sentence = re.sub(r'[^\x00-\x7F]+',' ', sentence).replace('\n',' ')
			s = {'text': sentence.encode('utf-8'), 'aspectTerms':'', 'aspectCategories':''}
			e = dict_to_xml('sentence',	 s)
			e.set('id',str(ind))
			writ = minidom.parseString(ET.tostring(e)).toprettyxml(indent="    ")
			f.write(writ)
		f.close()
